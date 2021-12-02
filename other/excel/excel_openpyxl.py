import openpyxl


# 利用openpyxl读写excel，注意这里只能是xlsx类型的excel
work_book = openpyxl.load_workbook('data_openpyxl.xlsx')
work_sheet = work_book.get_sheet_by_name('Sheet1')
row3 = [item.value for item in list(work_sheet.rows)[2]]
print(row3)

col3 = [item.value for item in list(work_sheet.columns)[2]]
print(col3)

cell_2_3 = work_sheet.cell(row=2, column=3).value
print(cell_2_3)

max_row = work_sheet.max_row
print(max_row)


work_book_write = openpyxl.Workbook()
write_sheet = work_book_write.active
write_sheet['A1'] = 'hi,wwu'
work_book_write.save('new.xlsx')


